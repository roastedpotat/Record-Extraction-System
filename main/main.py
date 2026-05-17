import time
start_time = time.perf_counter()
import os
from openpyxl import load_workbook
from llama_cpp import Llama
from concurrent.futures import ThreadPoolExecutor
from LLMFunctions.LlamaResponse import llm_response
from DriveFunctions.Google import Create_Service
from DriveFunctions.FolderDive import get_ParentFolderId, Search_Folder, get_folder_name
from DriveFunctions.GetImages import get_image_bytes
from OCRFunctions.TextRecognition import Text_from_images, Record_Grouping_with_Dates, new_edit_image
from OCRFunctions.OCRCleaner import master_clean_ocr, remove_sidebar_noise
from OCRFunctions.ICDManual import find_primary_icd
from ExcelFunctions.OpenExcel import load_in_file, get_column_names, add_data_to_excel
from ExcelFunctions.DataAddition import specific_id_row, row_num_checker, check_row_data, data_per_row
from JSONExtract.JSONtoPy import extract_json
from JSONExtract.CheckPointSystem import check_progress, save_progress
os.environ['PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK'] = 'True'
from paddleocr import PaddleOCR
import paddle

CLIENT_FILE = 'credentials.json'
API_NAME = 'drive'
API_VER = 'v3'
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']

filepath = r"g:\ER Visits(Above 60) - Jan 2020 to Jun 2025.xlsx"

sheet_name= "Visit Recs"

file = load_in_file(filepath= filepath, sheet_name= sheet_name)

column_names = get_column_names(file)

wb = load_workbook(filename= filepath)
ws = wb[sheet_name]

checkpointfile = "checkpoint.json"

system_content = '''You are a strict medical data extractor. Extract only 
                  what is explicitly present in the text. Never invent or 
                  infer data. If a field is not found, use null.'''

prompt = '''### EXTRACTION RULES:
1. **primary_icd:** You will be given the primary ICD code directly. Find its description from the "Provisional Diagnosis" table and return as "Code - Description".
2. **other_icd:** From the "Provisional Diagnosis" table, list all ICD codes that do NOT have "Yes" 
after them anywhere in the table. The table may contain duplicate header rows — treat all rows as part 
of one unified table. Return as comma-separated ICD codes only, no descriptions.
3. **chief_complaint:** Find the FIRST occurrence of "Chief Complaint & History of Present Illness" or "Chief Complaint and History of Present Illness". Extract only the text that comes after "Triage Category: X" after that heading. Stop before "Past History" or any other section heading.
4. **date_of_birth:** Find the text "DOB | Age | Gender:". Extract the value immediately following the label but before the first vertical bar (|). Ensure the date is formatted as DD/MM/YYYY.
5. **nationality:** Find "Nationality:" and return only the nationality value.
6. **Vitals:** Extract from the "Nursing Assessment" section only, specifically the first row of values under the labels "Temperature", "Pulse", "Respiratory", "BP", "O2SAT". These labels and their values may appear on separate lines. Do NOT wrap values in curly braces or any other characters — return plain values only.
- **temperature_celsius:** Value under "Temperature". Strip °C or *C. Return numeric only.
- **pulse_min:** Value under "Pulse". Strip "/min". Return numeric only.
- **respiratory_min:** Value under "Respiratory". Strip "/min". Return numeric only.
- **bp_mmhg:** Value under "BP". Return in format "120/80", strip "mm/Hg".
- **o2_sat:** Value under "O2SAT". 
If no number is directly present, find the first numeric value before or after a /% /symbol in the Nursing Assessment section. Return numeric only.If a vital is not found, return null.
7. **visit_date / visit_time:** Find the FIRST occurrence of "Visit Date:" followed by DD/MM/YYYY then a 4-digit time. Split into separate fields. If the date separator is missing (e.g. "08/08 2020"), reconstruct it as DD/MM/YYYY.
8. **disposition_date / disposition_time:** Go to the absolute end of the provided text. Search backwards from the very last character until you hit the first date and time pair. This value is likely the discharge/disposition time, from this enter the DD/MM/YYYY for date. Ignore all earlier timestamps found in the Triage or Vitals sections. Ensure the time you extract is specifically the one associated with the end of the patient's visit.
9. **pain_scale_score:** Find the FIRST occurrence of "Numerical(X)". Return only the integer X. Do not confuse with GCS.
10. **gcs:** Find the GCS value in the "Nursing Assessment" vitals section. Format is typically "XX/15". Return only the numerator as a plain integer. Do not return the full fraction.
11. **triage_category:** Find the FIRST occurrence of "Triage Category:" and return only the number after it.
12. **past_history:** Find the FIRST occurrence of "Past History:" or "Past Medical History". Extract only the text directly after it. Stop before "Travel History:" or any sidebar navigation text such as "HEMODIALYSIS", "LAB REPORTS", "OTHER DOCUMENTS", "ASSESSMENT/RE-ASSESSMENT". Do not include the label itself.
13. **occupation:** Find the FIRST occurrence of "Occupation:" and extract only the text directly after it.
14. **marital_status:** Find "Marital Status:" and return only the value after it.
15. **advice_health_education:** Find the FIRST occurrence of "Advice & Health Education:" and return only the text immediately after it. Stop at "Education Given To:" or any other label.
16. **condition_at_disposition:** Find "Condition at the time of Disposition:" or "Condition at the tine of Disposition:" and return only the value after it.
17. **disposition_type:** Find "Disposition Type:" and return only the value after it.
18. **remarks:** If remarks text is found after the Provisional Diagnosis table and before the Medication Order table, extract it and correct only obvious OCR spelling errors in individual words while preserving the original sentence structure and word order exactly. Otherwise return null.
19. **travel_history:** Find the FIRST occurrence of "Travel History:". Extract only the text directly after it. Return a value only if it describes a place, country, or region. If the value is a date, month-year, noise, or any non-place text, return null.
20. **current_medication:** Find the FIRST occurrence of "Current Medication". Extract only text directly after it. Stop at "Medical Prescription" or "Medication Order".
21. **psychosocial:** Find the FIRST occurrence of "Psychosocial:" and extract only the text after it. Do not include occupation.
22. **disease_grouping:** Unless specified, leave as null.
'''

paddle.set_flags({
"FLAGS_fraction_of_cpu_memory_to_use": 1.0,
"FLAGS_allocator_strategy": "naive_best_fit", # Optimize memory fragmentation
})

ocr_engine = PaddleOCR(
    use_doc_orientation_classify=False,
    use_doc_unwarping=False,
    use_textline_orientation= False,
    text_recognition_batch_size=2, 
    text_det_limit_side_len= 2000,
    text_det_limit_type= 'max',
    text_det_box_thresh= 0.5,
    text_det_thresh= 0.5,
    text_det_unclip_ratio=1.6, #Best results with 1.6
    lang='en',
    device= 'cpu',
    enable_mkldnn=True,
    mkldnn_cache_capacity=10,
    cpu_threads=4
    )

llm_engine = Llama(model_path= r"g:\models\qwen2.5-7b-instruct-q5_0-00001-of-00002.gguf",
                chat_format= "chatml",
                flash_attn= True,
                n_gpu_layers=-1,
                temperature= 0.2,
                seed= 1337,
                n_ctx= 8192,
                verbose= False)

service = Create_Service(CLIENT_FILE, API_NAME, API_VER, SCOPES)

parent_folder_id = get_ParentFolderId(body= service, 
                                      text= 'Test UploadFolder')

subfolder_ids = Search_Folder(body= service, parent_id= parent_folder_id)

for folder in subfolder_ids:
    comp_id = get_folder_name(body= service, folder_id= folder)

    checkpoint = check_progress(checkpointfile)

    if checkpoint >= int(comp_id):
        continue

    rows= specific_id_row(dataframe= file, specific_id= comp_id)

    image_ids = Search_Folder(body= service, parent_id= folder)

    image_bytes = get_image_bytes(image_file_ids= image_ids)

    readable_img = []
    with ThreadPoolExecutor(max_workers=16) as Executor:
        if image_bytes:
            readable_img = list(Executor.map(new_edit_image, image_bytes))
    
    texts = list(Text_from_images(ocr= ocr_engine, readable_list= readable_img))

    print("OCR done!")
    record = Record_Grouping_with_Dates(texts= texts)
    row_count_check = row_num_checker(rows= rows, total_records= record)
    row_data_check = check_row_data(rows= rows)
    if not row_data_check or row_count_check:
        if row_count_check > 0:
            new_row = max(row_data_check) + 3
            ws.insert_rows(new_row, row_count_check)
            copy_row_data = [cell.value for cell in ws[(max(row_data_check)+2)]]
        for _ in range(row_count_check):
            row_data_check.append(max(row_data_check) + 1)
        new_record = data_per_row(records= record, df= file, comp_id= comp_id)
        if not new_record:
            continue
        print("Cleaning Started!")
        clean_records = master_clean_ocr(records_dict= new_record)
        clean_noise_records = remove_sidebar_noise(records= clean_records)
        primary_icd = find_primary_icd(clean_noise_records)

        print("LLM processing...")
        llm_step = list(llm_response(llm= llm_engine, 
                                     clean_records= clean_noise_records, 
                                     column_names= column_names, 
                                     prompt= prompt, 
                                     system_content= system_content,
                                     primary_icd= primary_icd))
        for i,record in enumerate(llm_step):
            data = extract_json(record)
            print(f"\n\nLLM Output for record {i+1}:\n{data}\n\n")
            add_data_to_excel(data= data, 
                              ws= ws,
                              starting_column= 6, 
                              row_num= row_data_check[i],
                              initial_data= copy_row_data 
                              if row_count_check > 0 else None)
    save_progress(checkpointfile, comp_id= comp_id)
    wb.save(filename= filepath)

end_time = time.perf_counter()
elapsed_time = end_time - start_time
print(f"Took a total of {elapsed_time:.1f} seconds")